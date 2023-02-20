#!/usr/bin/python3
import os
import sys
import datetime
from copy import copy
termwidth = os.get_terminal_size()[0]
print("Welcome to Sukriti's Attendance Synchronization System (SASS)!".center(termwidth, '-'),"\n\n")

infile = ""
outfile = ""
interactive = False

for i in range (0,len(sys.argv)):
    match sys.argv[i]:
        case "-i":
            infile = sys.argv[i+1]
        case "-o":
            outfile = sys.argv[i+1]
        case "-I" | "--interactive":
            interactive = True

def showfileoptions(flist: list):
    for i in range(0,len(flist)):
        print("{}) {}".format(i+1,flist[i]))

def changedir():
    ch = input("\nPlease enter your desired path or press enter to enter interactive mode: ")
    if len(ch):
        try:
            os.chdir(ch)
            return
        except(FileNotFoundError):
            input("That wasn't a valid path. Press enter to continue.")
            return changedir()
    while True:
        print("\n------------------\nCurrent folder: ",os.getcwd())
        dirs = []
        for dir in os.scandir():
            if dir.is_dir():
                dirs.append(dir)
        for dir in range(0,len(dirs)):
            print("{}) {}".format(dir+i, dirs[dir].name))
        print("0) Go up one directory")
        print("\nIf you're done, press enter")
        
        choice = input("\n select an option: ")
        try:
            choice = int(choice)
        except(ValueError):
            if len(choice):
                print("That wasn't a number!")
                continue
            return
        if choice == 0:
            os.chdir("..")
        else:
            try:
                os.chdir(dirs[choice-1])
            except(IndexError):
                print("That wasn't an option! retrying")
                

def findfiles(name: str):
    print("Current directory: ",os.getcwd())
    xlfiles = []
    for f in os.scandir():
        f = f.name
        if f.count(".xlsx"):
            xlfiles.append(f)
    if len(xlfiles):
        print("Found {} excel files in current directory".format(len(xlfiles)))
        showfileoptions(xlfiles)
        print("0) Choose a different folder")
        try:
            choice = int(input("Please select the number of the {}: ".format(name)))
        except(ValueError):
            print("That wasn't a number you silly goose! \n")
            return findfiles(name)
        if choice == 0:
            changedir()
            return findfiles(name)
        else:
            return xlfiles[choice-1]
    else:
        print("I can't find any excel files in this folder".center(12,"!"))
        input("Press enter to continue...")
        changedir()
        return findfiles(name)
        
             

if ((not len(infile) or not len(outfile)) and not interactive):
    print("\n Bad command invocation. Printing Help text...\n")
    print("Usage:\npython3 sass.py -i [input.xlsx] -o [output.xlsx]")
    print("Options:")
    print("-i                   specify input excel/google sheets file")
    print("-o                   specify target / output excel file")
    print("-I --interactve      run in interactive file choosing mode")
    exit()

if interactive:
    print("\n Starting in interactive mode.\n")
    infile = findfiles("Downloaded Google Form")
    print("\n------\n")
    outfile = findfiles("Main excel sheet")
    print("out: {} in: {}".format(outfile,infile)) 
try:
    import openpyxl as op
except(ModuleNotFoundError):
    print("Excel support library not installed")
    if input("Install it now? (y/N) ").lower() == "y":
        os.system("pip3 install openpyxl")
        import openpyxl as op
    else:
        print("Fatal error! Please install openpyxl manually and try again.")
        exit()

print("\nThe output sheet may have formulas. I'll try my best to update them automatically, but this might have unexpected behavior. Alternativley, I can replace them with my own calculations.")
internalCalc = (input("Override Excel Formulas? (y/N) ").lower() == "y")

class Attendee:
    fname: str
    lname: str
    year: int
    email: str
    time: datetime.datetime

def read_file():
    attendees = []
    wb = op.load_workbook(filename=infile)
    sheet = wb.active
    labels = tuple(sheet.rows)[0]
    # assign default positions of data
    tsindex = 0
    fnindex = 1
    lnindex = 2
    yrindex = 3
    emindex = 4
    # auto-assign indexes based on first row
    for lbi in range(0,len(labels)):
        try:
            match labels[lbi].value.lower():
                case "timestamp":
                    tsindex = lbi
                case "first name":
                    fnindex = lbi
                case "last name":
                    lnindex = lbi
                case "grad year":
                    yrindex = lbi
                case "wpi email":
                    emindex = lbi
        except(ValueError, AttributeError):
            pass

    for row in sheet.iter_rows(min_row=2, values_only=True):
        attendees.append(Attendee())
        attendees[-1].fname = row[fnindex]
        attendees[-1].lname = row[lnindex]
        attendees[-1].time = row[tsindex]
        attendees[-1].year = int(row[yrindex])
        attendees[-1].email = row[emindex]
    print("Read {} attendees from input sheet.".format(len(attendees)))
    return attendees

class Event:
    def __init__(self, name, col):
        self.name = name
        self.col = col

def dateformatter(indate):
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    try:
        out = "{}-{}".format(indate.day, months[indate.month-1])
    except(AttributeError):
        out = input("Automated date system failed. Please input date of manually (MMM-DD): ")
    return out

# i'm writing this at 11 PM, it needs to be done by tomorrow, get ready for the worst search function you've ever seen in your life
def write_file():
    import shutil
    shutil.copyfile(outfile, outfile+".bak")
    print("Created backup of main Excel File as {}.bak\n".format(outfile))
    wb = op.load_workbook(filename=outfile)
    sheet = wb.active
    labels = tuple(sheet.rows)[0]
    events = []
    people = read_file()
    event = Event("default",0)
    yrindex = 0
    memberindex = 1
    termattendanceindex = 3
    for i in range(0, len(labels)):
        match labels[i].value:
            case "Event": # It's not intuitive but it's how the sheet is layed out
                memberindex = i
            case "":
                continue
            case _:
                if labels[i].value is not None:
                    events.append(Event(labels[i].value, i))
    termattendanceindex = events[-1].col+1
    if len(events):
        print("I found the following Events:".center(termwidth))
        for i in range(0, len(events)):
            print("{}) {}".format(i+1,events[i].name))
        print("\n0) Add new Event")
        try:
            choice = int(input("\n Please select an option: "))
            if choice:
                event = events[choice-1]
            else:
                sheet.insert_cols(events[-1].col+2)
                event = Event(input("Please enter a new event name: "),events[-1].col+1)
                labels = tuple(sheet.rows)[0]
                labels[event.col].value = event.name
                labels[event.col].fill = copy(sheet.cell(1,1).fill)
                dates = tuple(sheet.rows)[1]
                dates[event.col].value = dateformatter(people[0].time)
                dates[event.col].fill = copy(sheet.cell(1,1).fill)
        except(ValueError, IndexError):
            print("Bad input! Please enter a number!")
            write_file()
            return
    else:
        "No existing events were detected."
        event = Event(input("Please enter a new event name: "),memberindex+1)
        sheet.insert_cols(memberindex+1)
        labels = tuple(sheet.rows)[0]
        labels[event.col].value = event.name
        labels[event.col].fill = copy(sheet.cell(1,1).fill)
        dates = tuple(sheet.rows)[1]
        dates[event.col].value = dateformatter(people[0].time)
        dates[event.col].fill.bgColor = copy(sheet.cell(1,1).fill)
    for row in sheet.iter_rows(min_row=3, values_only=False):
        try:
            name = row[memberindex].value.lower()
        except(AttributeError):
            continue
        try:
            year = int(row[yrindex].value)
        except(TypeError):
            year = ""
        for person in people:
            if name == (person.fname.lower() + " " + person.lname.lower()):
                if row[event.col].value is None:
                    row[event.col].value = 1
                else:
                    row[event.col].value += 1
                people.remove(person)
                # Recalculate column indexes
                labels = tuple(sheet.rows)[0]
                for i in range(0, len(labels)):
                    match labels[i].value:
                        case "Event": # It's not intuitive but it's how the sheet is layed out
                            memberindex = i
                        case "":
                            continue
                termattendanceindex = events[-1].col+1
                if internalCalc:
                    tally = 0
                    for ev in events+[event]:
                        try:
                            tally += int(row[ev.col].value)
                        except(ValueError):
                            print("Internal Calulation failed. Not updating counts for {}".format(person.fname))
                            break
                    row[termattendanceindex+1].value = tally
                else:
                    formulae = []
                    for cell in row:
                        if type(cell.value) is str:
                            if cell.value.count("="):
                                formulae.append(cell.column-1)
                    if len(formulae) == 0:
                        print("did not detect any formulas, skipping update process")
                        break
                    oldcol = ""
                    newcol = ""
                    for chara in row[event.col-1].coordinate:
                        if chara.isalpha():
                            oldcol += chara
                    for chara in row[event.col].coordinate:
                        if chara.isalpha():
                            newcol += chara
                    for formula in formulae:
                        try:
                            row[formula].value = row[formula].value.replace(oldcol, newcol)
                        except(AttributeError):
                            print("tried to access cell {} whose type is {}".format(row[formula].coordinate, type(row[formula].value)))
                break
    if len(people):
        print("The following {} attendees were listed on the form but not present in the main sheet:")
        for i in people:
            print("* ", i.fname, i.lname)
        if input("Add attendees to main sheet? (Y/n)").lower() != 'n':
            for i in people:
                personName = i.fname + " " + i.lname
                sheet.append({yrindex+1: i.year, memberindex+1: personName, event.col+1: 1})
                newrow = tuple(sheet.rows)[-1]
                for cell in range(memberindex+1, event.col):
                    newrow[cell].value = 0
                firstcell = newrow[memberindex+1].coordinate
                lastcell = newrow[event.col].coordinate
                newrow[termattendanceindex+1].value = "=SUM({}:{})".format(firstcell,lastcell)
    wb.save(outfile)
                
write_file()
print("done.")
